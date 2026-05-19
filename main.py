"""
Sincronizador INMET API → banco de dados.

Chave de identificação de um registro: (codigo, data, hora_utc)
  - data     : tipo DATE (PostgreSQL)
  - hora_utc : string "HHMM" (ex: "1800")

Lógica:
  - Registro existe no banco → atualiza campos divergentes
  - Registro não existe      → insere
  - Registro da API com todos os campos nulos → ignorado
"""

import sys
import csv
import json
import time
import logging
import argparse
import requests
import psycopg
import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any
from datetime import datetime, date, timezone, timedelta
from pathlib import Path

TZ_MINUS4 = timezone(timedelta(hours=-4))

from config import DB_CONNSTR, API_TOKEN, API_BASE_URL, DB_TABLE, FIELD_MAP

SLEEP_BETWEEN_STATIONS = 2   # segundos entre estações no modo lista
API_MAX_RETRIES        = 3   # tentativas por estação
API_RETRY_BACKOFF      = 5   # segundos base entre tentativas (multiplicado pela tentativa)


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

log = logging.getLogger("inmet_sync")


def setup_logging() -> Path:
    log_dir = Path(__file__).parent / "logs"
    log_dir.mkdir(exist_ok=True)

    log_file = log_dir / f"{datetime.now().strftime('%Y-%m-%d')}.log"
    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)-8s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt)

    ch = logging.StreamHandler()
    ch.setFormatter(fmt)

    log.setLevel(logging.DEBUG)
    log.addHandler(fh)
    log.addHandler(ch)
    log.info("Log iniciado → %s", log_file)
    return log_dir


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

def fetch_api(data_ini: str, data_fim: str, codigo: str) -> list[dict]:
    """Retorna lista de registros horários da API para o período e estação."""
    safe_url = f"{API_BASE_URL}/{data_ini}/{data_fim}/{codigo}/***"
    log.info("Consultando API | estação=%s  período=%s → %s  url=%s", codigo, data_ini, data_fim, safe_url)

    url = f"{API_BASE_URL}/{data_ini}/{data_fim}/{codigo}/{API_TOKEN}"

    last_exc: Exception = RuntimeError("Sem resposta da API")
    for tentativa in range(1, API_MAX_RETRIES + 1):
        try:
            resp = requests.get(url, timeout=30)
            log.debug("Resposta HTTP %s | %s bytes", resp.status_code, len(resp.content))
            resp.raise_for_status()
            data = resp.json()
            log.info("API retornou %d registros.", len(data))
            return data
        except requests.HTTPError as e:
            log.error("API retornou erro HTTP %s: %s", e.response.status_code, e)
            raise
        except Exception as e:
            last_exc = e
            wait = API_RETRY_BACKOFF * tentativa
            if tentativa < API_MAX_RETRIES:
                log.warning(
                    "Falha na tentativa %d/%d para estação %s: %s — aguardando %ds antes de tentar novamente.",
                    tentativa, API_MAX_RETRIES, codigo, e, wait,
                )
                time.sleep(wait)
            else:
                log.error(
                    "Todas as %d tentativas falharam para estação %s: %s",
                    API_MAX_RETRIES, codigo, e,
                )

    raise last_exc


def is_null_record(api_rec: dict) -> bool:
    """Retorna True se todos os campos meteorológicos do registro são nulos."""
    result = all(api_rec.get(k) is None for k in FIELD_MAP)
    if result:
        log.debug(
            "Registro ignorado (todos os campos nulos) | estação=%s  data=%s  hora=%s",
            api_rec.get("CD_ESTACAO"), api_rec.get("DT_MEDICAO"), api_rec.get("HR_MEDICAO"),
        )
    return result


# ---------------------------------------------------------------------------
# Conversão de formatos
# ---------------------------------------------------------------------------

def api_date_to_db(api_date: str):
    """Converte "YYYY-MM-DD" (API) → objeto date do Python (coluna DATE no banco)."""
    return datetime.strptime(api_date, "%Y-%m-%d").date()


def make_id(api_date: str, hora_utc: str, codigo: str) -> str:
    """Gera o id_dado_inmet no padrão YYYYMMDDHHMM_CODIGO (ex: 202605071300_A709)."""
    date_part = datetime.strptime(api_date, "%Y-%m-%d").strftime("%Y%m%d")
    return f"{date_part}{hora_utc}_{codigo}"


def api_rec_to_db_row(api_rec: dict) -> dict:
    """
    Converte um registro da API para o dicionário com os nomes das colunas
    do banco. Inclui todos os campos necessários para INSERT.
    """
    data_db      = api_date_to_db(api_rec["DT_MEDICAO"])
    hora_utc_str = api_rec["HR_MEDICAO"]      # "1400" — usado no id e na chave
    hora_utc_int = int(hora_utc_str)          # 1400  — tipo real da coluna no banco
    codigo       = api_rec["CD_ESTACAO"]
    lat          = float(api_rec["VL_LATITUDE"])
    lon          = float(api_rec["VL_LONGITUDE"])

    h = hora_utc_int // 100
    m = hora_utc_int % 100
    data_hora_obs_utc = datetime(data_db.year, data_db.month, data_db.day, h, m, 0,
                                 tzinfo=TZ_MINUS4)

    row = {
        "id_dado_inmet":     make_id(api_rec["DT_MEDICAO"], hora_utc_str, codigo),
        "data":              data_db,
        "hora_utc":          hora_utc_int,
        "data_hora_obs_utc": data_hora_obs_utc,
        "codigo":    codigo,
        "nome":      api_rec["DC_NOME"],
        "latitude":  lat,
        "longitude": lon,
        # geom calculada no SQL via ST_SetSRID(ST_MakePoint(lon, lat), 4326)
    }

    for api_field, db_col in FIELD_MAP.items():
        val = api_rec.get(api_field)
        row[db_col] = float(val) if val is not None else None

    log.debug(
        "Registro convertido | id=%s  data_hora=%s",
        row["id_dado_inmet"], data_hora_obs_utc,
    )
    return row


# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------

def get_db_records(conn, codigo: str, data_ini: str, data_fim: str) -> dict:
    """
    Busca os registros do banco para o período e estação.
    Retorna dict indexado por (data, hora_utc) → row como dict.

    data_ini / data_fim no formato "YYYY-MM-DD".
    """
    d_ini = datetime.strptime(data_ini, "%Y-%m-%d").date()
    d_fim = datetime.strptime(data_fim, "%Y-%m-%d").date()

    log.info("Consultando banco | tabela=%s  estação=%s  período=%s → %s", DB_TABLE, codigo, d_ini, d_fim)

    sql = f"""
        SELECT data, hora_utc, id_dado_inmet, data_hora_obs_utc,
               {', '.join(FIELD_MAP.values())}
        FROM {DB_TABLE}
        WHERE codigo = %s
          AND data BETWEEN %s AND %s
        ORDER BY data, hora_utc
    """

    result = {}
    with conn.cursor() as cur:
        cur.execute(sql, (codigo, d_ini, d_fim))
        cols = [desc[0] for desc in cur.description]
        for row in cur.fetchall():
            rec = dict(zip(cols, row))
            result[(rec["data"], rec["hora_utc"])] = rec

    log.info("Banco retornou %d registros para o período.", len(result))
    return result


def update_record(conn, row: dict, existing: dict) -> bool:
    """Atualiza apenas os campos que diferem entre API e banco."""
    diffs = {}

    # Campos meteorológicos
    for db_col in FIELD_MAP.values():
        api_val = row.get(db_col)
        db_val  = existing.get(db_col)
        api_f = float(api_val) if api_val is not None else None
        db_f  = float(db_val)  if db_val  is not None else None
        if api_f != db_f:
            diffs[db_col] = api_val

    # Campos de identificação/timestamp que podem estar NULL em registros antigos
    if existing.get("id_dado_inmet") is None:
        diffs["id_dado_inmet"] = row["id_dado_inmet"]

    if existing.get("data_hora_obs_utc") is None:
        diffs["data_hora_obs_utc"] = row["data_hora_obs_utc"]

    if not diffs:
        log.debug("Sem diferenças | id=%s", existing.get("id_dado_inmet") or row["id_dado_inmet"])
        return False

    log.debug("Diferenças encontradas | id=%s  campos=%s", row["id_dado_inmet"], list(diffs.keys()))

    set_clause = ", ".join(f"{col} = %s" for col in diffs)
    values = list(diffs.values()) + [row["codigo"], row["data"], int(row["hora_utc"])]
    sql = f"""
        UPDATE {DB_TABLE}
        SET {set_clause}
        WHERE codigo = %s AND data = %s AND hora_utc = %s
    """
    with conn.cursor() as cur:
        cur.execute(sql, values)
        if cur.rowcount == 0:
            log.warning(
                "UPDATE executado mas nenhuma linha afetada | estação=%s  data=%s  hora=%s",
                row["codigo"], row["data"], row["hora_utc"],
            )
            return False

    log.debug("UPDATE ok | id=%s", row["id_dado_inmet"])
    return True


def insert_record(conn, row: dict):
    """Insere um novo registro no banco."""
    log.debug("Inserindo | id=%s  data_hora=%s", row["id_dado_inmet"], row["data_hora_obs_utc"])

    cols      = list(row.keys()) + ["geom"]
    ph_values = ["%s"] * len(row) + ["ST_SetSRID(ST_MakePoint(%s, %s), 4326)"]
    values    = list(row.values()) + [row["longitude"], row["latitude"]]

    sql = f"INSERT INTO {DB_TABLE} ({', '.join(cols)}) VALUES ({', '.join(ph_values)}) ON CONFLICT DO NOTHING"
    with conn.cursor() as cur:
        cur.execute(sql, values)

    log.debug("INSERT ok | id=%s", row["id_dado_inmet"])


# ---------------------------------------------------------------------------
# Sincronização
# ---------------------------------------------------------------------------

def sync(conn, api_records: list[dict], codigo: str, data_ini: str, data_fim: str) -> dict:
    log.info("Iniciando sincronização | estação=%s  período=%s → %s", codigo, data_ini, data_fim)

    db_records = get_db_records(conn, codigo, data_ini, data_fim)

    inserted = 0
    updated  = 0
    skipped  = 0

    for api_rec in api_records:
        if is_null_record(api_rec):
            skipped += 1
            continue

        row = api_rec_to_db_row(api_rec)
        key = (row["data"], row["hora_utc"])

        if key in db_records:
            if update_record(conn, row, db_records[key]):
                updated += 1
                log.info("UPDATE | id=%s", row["id_dado_inmet"])
        else:
            insert_record(conn, row)
            inserted += 1
            log.info("INSERT | id=%s", row["id_dado_inmet"])

    conn.commit()
    log.info("Commit realizado.")

    log.info(
        "Sincronização concluída | inseridos=%d  atualizados=%d  ignorados(sem dados)=%d",
        inserted, updated, skipped,
    )
    return {"inseridos": inserted, "atualizados": updated, "ignorados": skipped}


# ---------------------------------------------------------------------------
# Entrada
# ---------------------------------------------------------------------------

CODIGOS_FILE = Path(__file__).parent / "codigos_estacoes_ms.txt"


def input_date(prompt: str) -> str:
    """Solicita uma data no formato YYYY-MM-DD com validação básica."""
    while True:
        val = input(prompt).strip()
        try:
            datetime.strptime(val, "%Y-%m-%d")
            return val
        except ValueError:
            log.warning("Data inválida informada: '%s'. Use YYYY-MM-DD.", val)
            print("  Formato inválido. Use YYYY-MM-DD (ex: 2026-05-07).")


def input_mode() -> str:
    """Solicita ao usuário o modo de operação. Retorna '1' ou '2'."""
    print("\nModo de operação:")
    print("  [1] Uma estação")
    print("  [2] Todas as estações da lista (codigos_estacoes_ms.txt)")
    while True:
        val = input("Escolha o modo [1/2]: ").strip()
        if val in ("1", "2"):
            return val
        print("  Opção inválida. Digite 1 ou 2.")


def load_codigos() -> list[str]:
    """Lê os códigos de estação do arquivo, ignorando linhas vazias e duplicatas."""
    if not CODIGOS_FILE.exists():
        log.error("Arquivo de estações não encontrado: %s", CODIGOS_FILE)
        sys.exit(1)

    codigos = []
    seen = set()
    with open(CODIGOS_FILE, encoding="utf-8") as f:
        for line in f:
            code = line.strip().upper()
            if code and code not in seen:
                codigos.append(code)
                seen.add(code)

    log.info("Arquivo de estações carregado | %d estações únicas em %s", len(codigos), CODIGOS_FILE.name)
    return codigos


def sync_estacao(conn, codigo: str, data_ini: str, data_fim: str) -> dict:
    """Busca a API e sincroniza o banco para uma única estação."""
    api_data = fetch_api(data_ini, data_fim, codigo)
    return sync(conn, api_data, codigo, data_ini, data_fim)


def _build_resumo(ts_inicio: datetime, modo: str, data_ini: str, data_fim: str, estacoes: list[dict]) -> dict:
    ts_fim  = datetime.now()
    duracao = round((ts_fim - ts_inicio).total_seconds())
    totais  = {
        "estacoes":      len(estacoes),
        "estacoes_ok":   sum(1 for e in estacoes if e["status"] == "ok"),
        "estacoes_erro": sum(1 for e in estacoes if e["status"] == "erro"),
        "inseridos":     sum(e.get("inseridos", 0) for e in estacoes),
        "atualizados":   sum(e.get("atualizados", 0) for e in estacoes),
        "ignorados":     sum(e.get("ignorados", 0) for e in estacoes),
    }
    return {
        "inicio":           ts_inicio.strftime("%Y-%m-%d %H:%M:%S"),
        "fim":              ts_fim.strftime("%Y-%m-%d %H:%M:%S"),
        "duracao_segundos": duracao,
        "modo":             modo,
        "tabela":           DB_TABLE,
        "data_ini":         data_ini,
        "data_fim":         data_fim,
        "totais":           totais,
        "estacoes":         estacoes,
    }


def _save_json(path: Path, resumo: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(resumo, f, ensure_ascii=False, indent=2)


def _save_csv(path: Path, resumo: dict) -> None:
    meta = [
        ("inicio",           resumo["inicio"]),
        ("fim",              resumo["fim"]),
        ("duracao_segundos", resumo["duracao_segundos"]),
        ("modo",             resumo["modo"]),
        ("tabela",           resumo["tabela"]),
        ("data_ini",         resumo["data_ini"]),
        ("data_fim",         resumo["data_fim"]),
    ]
    totais_rows = [(k, v) for k, v in resumo["totais"].items()]

    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)

        w.writerow(["# METADADOS", ""])
        w.writerows(meta)
        w.writerow([])

        w.writerow(["# TOTAIS", ""])
        w.writerows(totais_rows)
        w.writerow([])

        w.writerow(["# ESTAÇÕES"])
        w.writerow(["codigo", "status", "inseridos", "atualizados", "ignorados", "erro"])
        for e in resumo["estacoes"]:
            w.writerow([
                e["codigo"], e["status"],
                e.get("inseridos", 0), e.get("atualizados", 0), e.get("ignorados", 0),
                e.get("erro", ""),
            ])


def _xlsx_write_section(
    ws: Worksheet,
    title: str,
    rows: list[tuple[str, Any]],
    start_row: int,
    section_font: Font,
    section_fill: PatternFill,
) -> int:
    title_cell: Cell = ws.cell(start_row, 1)  # type: ignore[assignment]
    title_cell.value = title
    title_cell.font  = section_font
    title_cell.fill  = section_fill
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    start_row += 1
    for label, value in rows:
        label_cell: Cell = ws.cell(start_row, 1)  # type: ignore[assignment]
        value_cell: Cell = ws.cell(start_row, 2)  # type: ignore[assignment]
        label_cell.value = label
        label_cell.font  = Font(bold=True)
        value_cell.value = value
        start_row += 1
    return start_row + 1


def _save_xlsx(path: Path, resumo: dict) -> None:
    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2E75B6")
    section_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="404040")

    # --- Aba Resumo ---
    ws_resumo: Worksheet = wb.active  # type: ignore[assignment]
    ws_resumo.title = "Resumo"

    meta_rows: list[tuple[str, Any]] = [
        ("Início",             resumo["inicio"]),
        ("Fim",                resumo["fim"]),
        ("Duração (segundos)", resumo["duracao_segundos"]),
        ("Modo",               resumo["modo"]),
        ("Tabela",             resumo["tabela"]),
        ("Data inicial",       resumo["data_ini"]),
        ("Data final",         resumo["data_fim"]),
    ]
    totais_rows: list[tuple[str, Any]] = [
        (k.replace("_", " ").capitalize(), v) for k, v in resumo["totais"].items()
    ]

    next_row = _xlsx_write_section(ws_resumo, "METADADOS", meta_rows, 1, section_font, section_fill)
    _xlsx_write_section(ws_resumo, "TOTAIS", totais_rows, next_row, section_font, section_fill)

    ws_resumo.column_dimensions["A"].width = 28
    ws_resumo.column_dimensions["B"].width = 28

    # --- Aba Estações ---
    ws_est: Worksheet = wb.create_sheet("Estações")
    headers = ["Código", "Status", "Inseridos", "Atualizados", "Ignorados", "Erro"]
    for col, h in enumerate(headers, 1):
        cell: Cell = ws_est.cell(1, col)  # type: ignore[assignment]
        cell.value     = h
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    ok_fill   = PatternFill("solid", fgColor="E2EFDA")
    erro_fill = PatternFill("solid", fgColor="FFDDC1")

    for r, e in enumerate(resumo["estacoes"], 2):
        row_fill = ok_fill if e["status"] == "ok" else erro_fill
        values: list[Any] = [
            e["codigo"], e["status"],
            e.get("inseridos", 0), e.get("atualizados", 0), e.get("ignorados", 0),
            e.get("erro", ""),
        ]
        for col, val in enumerate(values, 1):
            c: Cell = ws_est.cell(r, col)  # type: ignore[assignment]
            c.value = val
            c.fill  = row_fill

    for col, width in zip("ABCDEF", [12, 10, 12, 14, 12, 60]):
        ws_est.column_dimensions[col].width = width

    wb.save(path)


def input_formato() -> str:
    """Solicita ao usuário o formato do arquivo de resumo."""
    print("\nFormato do resumo de execução:")
    print("  [1] JSON")
    print("  [2] CSV")
    print("  [3] XLSX")
    while True:
        val = input("Escolha o formato [1/2/3]: ").strip()
        if val in ("1", "2", "3"):
            return val
        print("  Opção inválida. Digite 1, 2 ou 3.")


def save_resumo(
    log_dir: Path,
    ts_inicio: datetime,
    modo: str,
    data_ini: str,
    data_fim: str,
    estacoes: list[dict],
    formato: str,
) -> None:
    resumo  = _build_resumo(ts_inicio, modo, data_ini, data_fim, estacoes)
    ext_map = {"1": "json", "2": "csv", "3": "xlsx"}
    ext     = ext_map[formato]
    path    = log_dir / f"{ts_inicio.strftime('%Y-%m-%d_%H-%M-%S')}_resumo.{ext}"

    if formato == "1":
        _save_json(path, resumo)
    elif formato == "2":
        _save_csv(path, resumo)
    else:
        _save_xlsx(path, resumo)

    log.info("Resumo salvo → %s", path)


FORMAT_MAP = {"json": "1", "csv": "2", "xlsx": "3"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sincronizador INMET → Banco de Dados",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--modo", choices=["estacao", "lista"],
        help="Modo de operação:\n  estacao — uma estação\n  lista   — todas do arquivo codigos_estacoes_ms.txt",
    )
    parser.add_argument("--codigo", help="Código da estação (obrigatório se --modo estacao)")
    parser.add_argument("--data-ini", dest="data_ini", help="Data inicial YYYY-MM-DD")
    parser.add_argument("--data-fim", dest="data_fim", help="Data final   YYYY-MM-DD")
    parser.add_argument(
        "--dias-atras", dest="dias_atras", type=int, default=3,
        help="Calcula data_ini = hoje - N dias, data_fim = hoje (padrão: 3).\nIgnorado se --data-ini/--data-fim forem informados.",
    )
    parser.add_argument(
        "--formato", choices=["json", "csv", "xlsx"], default="xlsx",
        help="Formato do arquivo de resumo (padrão: xlsx)",
    )
    return parser.parse_args()


def resolve_dates(args: argparse.Namespace) -> tuple[str, str]:
    """Retorna (data_ini, data_fim) a partir dos args, calculando por dias-atras se necessário."""
    if args.data_ini and args.data_fim:
        return args.data_ini, args.data_fim
    today    = date.today()
    data_fim = today.strftime("%Y-%m-%d")
    data_ini = (today - timedelta(days=args.dias_atras)).strftime("%Y-%m-%d")
    return data_ini, data_fim


def run(modo_label: str, data_ini: str, data_fim: str, formato: str, codigo_unico: str = "") -> None:
    """Executa a sincronização com os parâmetros já resolvidos (usado tanto pelo modo interativo quanto pelo CLI)."""
    log_dir   = setup_logging()
    ts_inicio = datetime.now()

    log.info("=" * 50)
    log.info("  Sincronizador INMET → Banco de Dados")
    log.info("=" * 50)
    log.info("Modo=%s  data_ini=%s  data_fim=%s  formato=%s", modo_label, data_ini, data_fim, formato)

    estacoes_resumo: list[dict] = []

    log.info("Conectando ao banco de dados...")
    with psycopg.connect(DB_CONNSTR) as conn:
        log.info("Conexão estabelecida com sucesso.")

        if modo_label == "unica_estacao":
            log.info("Estação informada: %s", codigo_unico)
            try:
                stats = sync_estacao(conn, codigo_unico, data_ini, data_fim)
                estacoes_resumo.append({"codigo": codigo_unico, "status": "ok", **stats})
            except Exception as e:
                log.error("Falha na estação %s: %s", codigo_unico, e)
                estacoes_resumo.append({"codigo": codigo_unico, "status": "erro", "erro": str(e),
                                        "inseridos": 0, "atualizados": 0, "ignorados": 0})
        else:
            codigos = load_codigos()
            total   = len(codigos)

            for i, codigo in enumerate(codigos, start=1):
                log.info("-" * 50)
                log.info("Processando estação %d/%d: %s", i, total, codigo)
                try:
                    stats = sync_estacao(conn, codigo, data_ini, data_fim)
                    estacoes_resumo.append({"codigo": codigo, "status": "ok", **stats})
                except Exception as e:
                    log.error("Falha na estação %s: %s — continuando...", codigo, e)
                    estacoes_resumo.append({"codigo": codigo, "status": "erro", "erro": str(e),
                                            "inseridos": 0, "atualizados": 0, "ignorados": 0})

                if i < total:
                    log.debug("Aguardando %ds antes da próxima estação...", SLEEP_BETWEEN_STATIONS)
                    time.sleep(SLEEP_BETWEEN_STATIONS)

            erros = [e["codigo"] for e in estacoes_resumo if e["status"] == "erro"]
            log.info("=" * 50)
            log.info("Lote concluído | total=%d  com erro=%d", total, len(erros))
            if erros:
                log.warning("Estações com erro: %s", ", ".join(erros))

    save_resumo(log_dir, ts_inicio, modo_label, data_ini, data_fim, estacoes_resumo, FORMAT_MAP[formato])
    log.info("Processo finalizado.")


def main() -> None:
    args = parse_args()

    # ── Modo automático (CLI): todos os parâmetros vieram por argumento ──
    if args.modo:
        if args.modo == "estacao" and not args.codigo:
            print("Erro: --codigo é obrigatório quando --modo estacao.")
            sys.exit(1)
        data_ini, data_fim = resolve_dates(args)
        modo_label = "unica_estacao" if args.modo == "estacao" else "lista"
        run(modo_label, data_ini, data_fim, args.formato, codigo_unico=args.codigo or "")

    # ── Modo interativo: sem argumentos, pede input ao usuário ──
    else:
        setup_logging()
        modo       = input_mode()
        modo_label = "unica_estacao" if modo == "1" else "lista"
        data_ini   = input_date("Data inicial (YYYY-MM-DD): ")
        data_fim   = input_date("Data final   (YYYY-MM-DD): ")
        formato    = input_formato()
        codigo_unico = ""
        if modo == "1":
            codigo_unico = input("Código da estação (ex: A702): ").strip().upper()
        run(modo_label, data_ini, data_fim, {"1": "json", "2": "csv", "3": "xlsx"}[formato], codigo_unico)


if __name__ == "__main__":
    main()
